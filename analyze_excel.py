
import openpyxl

# Carregar a planilha
wb = openpyxl.load_workbook('4.MOYA E LARA SOCIEDADE DE ADVOGADOS_PRÉVIA BASE ATIVA _ABRIL_disp_24_04_2025.xlsx')
sheet = wb.active

# Imprimir dimensões
print(f'Colunas: {sheet.max_column}')
print(f'Linhas: {sheet.max_row}')

# Imprimir cabeçalhos
print('\nCabeçalhos das colunas:')
for i in range(1, sheet.max_column + 1):
    value = sheet.cell(row=1, column=i).value
    print(f'{i}. {value}')

# Analisar tipos de dados (amostra das primeiras 5 linhas)
print('\nTipos de dados (primeiras 5 linhas):')
for col in range(1, sheet.max_column + 1):
    header = sheet.cell(row=1, column=col).value
    print(f'\nColuna {col}: {header}')
    print('Amostra de valores:')
    for row in range(2, min(7, sheet.max_row + 1)):
        value = sheet.cell(row=row, column=col).value
        value_type = type(value).__name__
        print(f'  Linha {row-1}: {value} (tipo: {value_type})')

# Contagem de células vazias por coluna
print('\nAnálise de células vazias:')
empty_cells = [0] * sheet.max_column
for col in range(1, sheet.max_column + 1):
    for row in range(2, sheet.max_row + 1):
        if sheet.cell(row=row, column=col).value is None:
            empty_cells[col-1] += 1

for col in range(1, sheet.max_column + 1):
    header = sheet.cell(row=1, column=col).value
    empty_count = empty_cells[col-1]
    total_rows = sheet.max_row - 1  # Excluindo o cabeçalho
    empty_percentage = (empty_count / total_rows) * 100
    print(f'Coluna {col} ({header}): {empty_count} células vazias ({empty_percentage:.2f}%)')
