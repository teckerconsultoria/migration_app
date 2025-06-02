import pandas as pd
import numpy as np
from datetime import datetime
from pathlib import Path
import logging
import json
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
import warnings
warnings.filterwarnings('ignore')

# Configurar logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('migration.log'),
        logging.StreamHandler()
    ]
)

class GCPJDataMigration:
    """
    Sistema de migração de dados entre planilhas especializadas.
    Realiza a migração considerando a completude dos dados e tratando valores ausentes.
    """
    
    def __init__(self):
        self.template_columns = [
            'CÓD. INTERNO', 'PROCESSO', 'PROCEDIMENTO', 'NOME PARTE CONTRÁRIA PRINCIPAL',
            'CPF/CNPJ', 'DEVEDOR SOLIDÁRIO 1', 'CPF 1', 'DEVEDOR SOLIDÁRIO 2', 'CPF 2',
            'ORGANIZAÇÃO CLIENTE', 'ESCRITÓRIO', 'TIPO DE OPERAÇÃO/CARTEIRA', 'OPERAÇÃO',
            'AGÊNCIA', 'CONTA', 'SEGMENTO DO CONTRATO', 'VARA', 'FORUM', 'COMARCA', 'UF',
            'ALERTA', 'PROVIDÊNCIA', 'MONITORAMENTO', 'PRAZO', 'PRAZO JUDICIAL', 'PLANO',
            'GESTOR', 'ÁREA RESPONSÁVEL PROCESSO (EX-DEJUR)', 'RESPONSÁVEL PROCESSO (EX-DEJUR)',
            'ÁREA RESPONSÁVEL PROCESSO [ESCRITÓRIO]', 'RESPONSÁVEL PROCESSO [ESCRITÓRIO]',
            'COMENTÁRIO (TICKET DE AJUIZAMENTO)'
        ]
        
        # Mapeamento detalhado de colunas
        self.column_mapping = {
            'CÓD. INTERNO': {'source': 'primary', 'field': 'GCPJ', 'type': 'string'},
            'PROCESSO': {'source': 'primary', 'field': 'PROCESSO', 'type': 'string'},
            'PROCEDIMENTO': {'source': 'primary', 'field': 'TIPO_ACAO', 'type': 'string'},
            'NOME PARTE CONTRÁRIA PRINCIPAL': {'source': 'primary', 'field': 'ENVOLVIDO', 'type': 'string'},
            'CPF/CNPJ': {'source': 'primary', 'field': 'CPF', 'type': 'string'},
            'DEVEDOR SOLIDÁRIO 1': {'source': 'juridico', 'field': 'AVALISTA 1', 'type': 'string'},
            'CPF 1': {'source': 'juridico', 'field': 'CPF/CNPJ Avalista 1', 'type': 'string'},
            'DEVEDOR SOLIDÁRIO 2': {'source': 'juridico', 'field': 'AVALISTA 2', 'type': 'string'},
            'CPF 2': {'source': 'juridico', 'field': 'CPF/CNPJ Avalista 2', 'type': 'string'},
            'ORGANIZAÇÃO CLIENTE': {'source': 'constant', 'value': 'BANCO BRADESCO S.A'},
            'ESCRITÓRIO': {'source': 'constant', 'value': 'MOYA E LARA'},
            'TIPO DE OPERAÇÃO/CARTEIRA': {'source': 'primary', 'field': 'CARTEIRA', 'type': 'string'},
            'AGÊNCIA': {'source': 'primary', 'field': 'AGENCIA', 'type': 'string'},
            'CONTA': {'source': 'primary', 'field': 'CONTA', 'type': 'string'},
            'SEGMENTO DO CONTRATO': {'source': 'secondary', 'field': 'TIPO', 'type': 'string'},
            'VARA': {'source': 'primary', 'field': 'ORGAO_JULGADOR', 'type': 'string'},
            'COMARCA': {'source': 'primary', 'field': 'COMARCA', 'type': 'string'},
            'UF': {'source': 'primary', 'field': 'UF', 'type': 'string'},
            'GESTOR': {'source': 'primary', 'field': 'GESTOR', 'type': 'string'},
            'RESPONSÁVEL PROCESSO [ESCRITÓRIO]': {'source': 'primary', 'field': 'ADVOGADO_CONTRATADO', 'type': 'string'}
        }
        
        self.migration_stats = {
            'total_gcpj': 0,
            'gcpj_migrated': 0,
            'gcpj_not_found': 0,
            'total_records': 0,
            'records_complete': 0,
            'records_partial': 0,
            'columns_filled': {},
            'errors': []
        }
        
    def load_data(self, gcpj_file, primary_file, secondary_file, juridico_file, template_file):
        """
        Carrega todos os arquivos necessários para a migração.
        """
        logging.info("Iniciando carregamento dos dados...")
        
        try:
            # Carregar lista de GCPJs
            self.gcpj_df = pd.read_csv(gcpj_file)
            self.gcpj_list = self.gcpj_df['GCPJ'].astype(str).tolist()
            self.migration_stats['total_gcpj'] = len(self.gcpj_list)
            logging.info(f"✓ {len(self.gcpj_list)} GCPJs carregados para migração")
            
            # Carregar planilhas
            self.primary_df = pd.read_excel(primary_file)
            self.primary_df['GCPJ'] = self.primary_df['GCPJ'].astype(str)
            logging.info(f"✓ Planilha primária carregada: {len(self.primary_df)} registros")
            
            self.secondary_df = pd.read_excel(secondary_file)
            self.secondary_df['GCPJ'] = self.secondary_df['GCPJ'].astype(str)
            logging.info(f"✓ Planilha secundária carregada: {len(self.secondary_df)} registros")
            
            self.juridico_df = pd.read_excel(juridico_file)
            self.juridico_df['GCPJ'] = self.juridico_df['GCPJ'].astype(str)
            logging.info(f"✓ Base jurídica carregada: {len(self.juridico_df)} registros")
            
            # Carregar template
            self.template_df = pd.read_excel(template_file, sheet_name='Sheet')
            logging.info(f"✓ Template carregado com {len(self.template_df.columns)} colunas")
            
            # Filtrar apenas GCPJs da lista
            self.primary_filtered = self.primary_df[self.primary_df['GCPJ'].isin(self.gcpj_list)]
            self.secondary_filtered = self.secondary_df[self.secondary_df['GCPJ'].isin(self.gcpj_list)]
            self.juridico_filtered = self.juridico_df[self.juridico_df['GCPJ'].isin(self.gcpj_list)]
            
            logging.info(f"Dados filtrados: {len(self.primary_filtered)} registros primários, "
                        f"{len(self.secondary_filtered)} secundários, "
                        f"{len(self.juridico_filtered)} jurídicos")
            
        except Exception as e:
            logging.error(f"Erro ao carregar dados: {str(e)}")
            raise
    
    def get_value_for_column(self, gcpj, contrato, column_name):
        """
        Obtém o valor para uma coluna específica baseado no mapeamento.
        """
        mapping = self.column_mapping.get(column_name)
        if not mapping:
            return None
        
        source = mapping['source']
        
        if source == 'constant':
            return mapping['value']
        
        elif source == 'primary':
            field = mapping['field']
            # Buscar no dataframe primário filtrado por GCPJ e contrato
            data = self.primary_filtered[
                (self.primary_filtered['GCPJ'] == gcpj) & 
                (self.primary_filtered['CONTRATO'] == contrato)
            ]
            if not data.empty and field in data.columns:
                value = data.iloc[0][field]
                return value if pd.notna(value) else None
        
        elif source == 'secondary':
            field = mapping['field']
            # Buscar no dataframe secundário por GCPJ
            data = self.secondary_filtered[self.secondary_filtered['GCPJ'] == gcpj]
            if not data.empty and field in data.columns:
                value = data.iloc[0][field]
                return value if pd.notna(value) else None
        
        elif source == 'juridico':
            field = mapping['field']
            # Buscar no dataframe jurídico por GCPJ
            data = self.juridico_filtered[self.juridico_filtered['GCPJ'] == gcpj]
            if not data.empty and field in data.columns:
                value = data.iloc[0][field]
                return value if pd.notna(value) else None
        
        return None
    
    def migrate_data(self):
        """
        Realiza a migração dos dados.
        """
        logging.info("\nIniciando migração dos dados...")
        
        migrated_data = []
        gcpj_processed = set()
        
        # Processar cada GCPJ
        for gcpj in self.gcpj_list:
            # Verificar se o GCPJ tem dados na base primária
            primary_data = self.primary_filtered[self.primary_filtered['GCPJ'] == gcpj]
            
            if primary_data.empty:
                self.migration_stats['gcpj_not_found'] += 1
                logging.warning(f"GCPJ {gcpj} não encontrado na base primária")
                continue
            
            gcpj_processed.add(gcpj)
            
            # Para cada contrato do GCPJ
            for _, row in primary_data.iterrows():
                contrato = row.get('CONTRATO', '')
                record = {}
                filled_columns = 0
                total_expected_columns = len([k for k, v in self.column_mapping.items() if v])
                
                # Preencher cada coluna do template
                for column in self.template_columns:
                    value = self.get_value_for_column(gcpj, contrato, column)
                    record[column] = value
                    
                    # Contabilizar preenchimento
                    if value is not None:
                        filled_columns += 1
                        if column not in self.migration_stats['columns_filled']:
                            self.migration_stats['columns_filled'][column] = 0
                        self.migration_stats['columns_filled'][column] += 1
                
                # Classificar o registro
                if filled_columns == total_expected_columns:
                    self.migration_stats['records_complete'] += 1
                elif filled_columns > 0:
                    self.migration_stats['records_partial'] += 1
                
                migrated_data.append(record)
                self.migration_stats['total_records'] += 1
                
                if len(migrated_data) % 1000 == 0:
                    logging.info(f"Processados {len(migrated_data)} registros...")
        
        self.migration_stats['gcpj_migrated'] = len(gcpj_processed)
        self.migrated_df = pd.DataFrame(migrated_data, columns=self.template_columns)
        
        logging.info(f"\n✓ Migração concluída: {len(migrated_data)} registros processados")
        logging.info(f"  - GCPJs migrados: {self.migration_stats['gcpj_migrated']}")
        logging.info(f"  - GCPJs não encontrados: {self.migration_stats['gcpj_not_found']}")
        logging.info(f"  - Registros completos: {self.migration_stats['records_complete']}")
        logging.info(f"  - Registros parciais: {self.migration_stats['records_partial']}")
        
        return self.migrated_df
    
    def apply_data_formatting(self):
        """
        Aplica formatação aos dados migrados.
        """
        logging.info("\nAplicando formatação aos dados...")
        
        # Formatar CPF/CNPJ
        cpf_columns = ['CPF/CNPJ', 'CPF 1', 'CPF 2']
        for col in cpf_columns:
            if col in self.migrated_df.columns:
                self.migrated_df[col] = self.migrated_df[col].apply(self.format_cpf_cnpj)
        
        # Padronizar strings
        for col in self.migrated_df.columns:
            if self.migrated_df[col].dtype == 'object':
                self.migrated_df[col] = self.migrated_df[col].apply(
                    lambda x: str(x).strip().upper() if pd.notna(x) else None
                )
        
        logging.info("✓ Formatação aplicada")
    
    def format_cpf_cnpj(self, value):
        """
        Formata CPF ou CNPJ.
        """
        if pd.isna(value):
            return None
        
        # Remover caracteres não numéricos
        value = str(value).replace('.', '').replace('-', '').replace('/', '').strip()
        
        if len(value) == 11:  # CPF
            return f"{value[:3]}.{value[3:6]}.{value[6:9]}-{value[9:]}"
        elif len(value) == 14:  # CNPJ
            return f"{value[:2]}.{value[2:5]}.{value[5:8]}/{value[8:12]}-{value[12:]}"
        else:
            return value
    
    def generate_migration_report(self):
        """
        Gera relatório detalhado da migração.
        """
        logging.info("\nGerando relatório de migração...")
        
        report = {
            'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'summary': {
                'total_gcpj_requested': self.migration_stats['total_gcpj'],
                'gcpj_migrated': self.migration_stats['gcpj_migrated'],
                'gcpj_not_found': self.migration_stats['gcpj_not_found'],
                'total_records_generated': self.migration_stats['total_records'],
                'records_complete': self.migration_stats['records_complete'],
                'records_partial': self.migration_stats['records_partial']
            },
            'column_completeness': {},
            'data_quality_issues': []
        }
        
        # Calcular completude por coluna
        for column in self.template_columns:
            filled = self.migration_stats['columns_filled'].get(column, 0)
            total = self.migration_stats['total_records']
            percentage = (filled / total * 100) if total > 0 else 0
            
            report['column_completeness'][column] = {
                'filled': filled,
                'total': total,
                'percentage': round(percentage, 2)
            }
        
        # Identificar problemas de qualidade
        if self.migration_stats['gcpj_not_found'] > 0:
            report['data_quality_issues'].append({
                'issue': 'GCPJs não encontrados',
                'count': self.migration_stats['gcpj_not_found'],
                'severity': 'HIGH'
            })
        
        # Salvar relatório
        with open('migration_report.json', 'w', encoding='utf-8') as f:
            json.dump(report, f, indent=2, ensure_ascii=False)
        
        logging.info("✓ Relatório salvo em migration_report.json")
        
        return report
    
    def export_results(self, output_file='dados_migrados.xlsx'):
        """
        Exporta os dados migrados para Excel com formatação.
        """
        logging.info(f"\nExportando resultados para {output_file}...")
        
        # Criar cópia do template
        wb = load_workbook('templatebancobradescosa.xlsx')
        ws = wb['Sheet']
        
        # Limpar dados existentes (mantendo headers)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.value = None
        
        # Adicionar dados migrados
        for idx, row in self.migrated_df.iterrows():
            for col_idx, column in enumerate(self.template_columns):
                ws.cell(row=idx+2, column=col_idx+1, value=row[column])
        
        # Aplicar formatação condicional
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
        
        # Destacar células vazias
        for row in ws.iter_rows(min_row=2, max_row=len(self.migrated_df)+1):
            for cell in row:
                if cell.value is None:
                    cell.fill = yellow_fill
        
        # Salvar arquivo
        wb.save(output_file)
        logging.info(f"✓ Dados exportados para {output_file}")
        
        # Exportar também em CSV para análise
        self.migrated_df.to_csv('dados_migrados.csv', index=False, encoding='utf-8-sig')
        logging.info("✓ Dados também exportados para dados_migrados.csv")
    
    def generate_quality_dashboard(self):
        """
        Gera um dashboard HTML com indicadores de qualidade da migração.
        """
        logging.info("\nGerando dashboard de qualidade...")
        
        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <title>Dashboard de Qualidade - Migração GCPJ</title>
            <meta charset="utf-8">
            <style>
                body {{ font-family: Arial, sans-serif; margin: 20px; background-color: #f5f5f5; }}
                .container {{ max-width: 1200px; margin: 0 auto; }}
                .header {{ background-color: #1e88e5; color: white; padding: 20px; border-radius: 5px; }}
                .metric-card {{ background-color: white; padding: 20px; margin: 10px; border-radius: 5px; 
                              box-shadow: 0 2px 5px rgba(0,0,0,0.1); display: inline-block; min-width: 200px; }}
                .success {{ color: #4caf50; }}
                .warning {{ color: #ff9800; }}
                .error {{ color: #f44336; }}
                .progress-bar {{ width: 100%; height: 20px; background-color: #e0e0e0; border-radius: 10px; }}
                .progress-fill {{ height: 100%; background-color: #4caf50; border-radius: 10px; }}
                table {{ width: 100%; border-collapse: collapse; background-color: white; }}
                th, td {{ padding: 10px; text-align: left; border-bottom: 1px solid #ddd; }}
                th {{ background-color: #1e88e5; color: white; }}
            </style>
        </head>
        <body>
            <div class="container">
                <div class="header">
                    <h1>Dashboard de Qualidade - Migração GCPJ</h1>
                    <p>Executado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}</p>
                </div>
                
                <h2>Métricas Gerais</h2>
                <div>
                    <div class="metric-card">
                        <h3>GCPJs Processados</h3>
                        <p style="font-size: 36px; font-weight: bold;" class="success">
                            {self.migration_stats['gcpj_migrated']} / {self.migration_stats['total_gcpj']}
                        </p>
                        <p>{self.migration_stats['gcpj_migrated']/self.migration_stats['total_gcpj']*100:.1f}% de sucesso</p>
                    </div>
                    
                    <div class="metric-card">
                        <h3>Registros Gerados</h3>
                        <p style="font-size: 36px; font-weight: bold;">
                            {self.migration_stats['total_records']}
                        </p>
                    </div>
                    
                    <div class="metric-card">
                        <h3>Registros Completos</h3>
                        <p style="font-size: 36px; font-weight: bold;" class="success">
                            {self.migration_stats['records_complete']}
                        </p>
                        <p>{self.migration_stats['records_complete']/self.migration_stats['total_records']*100:.1f}% do total</p>
                    </div>
                    
                    <div class="metric-card">
                        <h3>Registros Parciais</h3>
                        <p style="font-size: 36px; font-weight: bold;" class="warning">
                            {self.migration_stats['records_partial']}
                        </p>
                        <p>{self.migration_stats['records_partial']/self.migration_stats['total_records']*100:.1f}% do total</p>
                    </div>
                </div>
                
                <h2>Completude por Coluna</h2>
                <table>
                    <tr>
                        <th>Coluna</th>
                        <th>Registros Preenchidos</th>
                        <th>Taxa de Preenchimento</th>
                        <th>Visualização</th>
                    </tr>
                    {self._generate_column_quality_rows()}
                </table>
            </div>
        </body>
        </html>
        """
        
        with open('dashboard_qualidade_migracao.html', 'w', encoding='utf-8') as f:
            f.write(html_content)
        
        logging.info("✓ Dashboard salvo em dashboard_qualidade_migracao.html")
    
    def _generate_column_quality_rows(self):
        """
        Gera as linhas da tabela de qualidade por coluna.
        """
        rows = []
        for column in self.template_columns:
            filled = self.migration_stats['columns_filled'].get(column, 0)
            total = self.migration_stats['total_records']
            percentage = (filled / total * 100) if total > 0 else 0
            
            # Definir cor baseado na porcentagem
            if percentage >= 80:
                color_class = 'success'
            elif percentage >= 50:
                color_class = 'warning'
            else:
                color_class = 'error'
            
            # Criar barra de progresso
            progress_bar = f'''
                <div class="progress-bar">
                    <div class="progress-fill" style="width: {percentage}%"></div>
                </div>
            '''
            
            rows.append(f"""
                <tr>
                    <td>{column}</td>
                    <td>{filled} / {total}</td>
                    <td class="{color_class}">{percentage:.1f}%</td>
                    <td>{progress_bar}</td>
                </tr>
            """)
        
        return '\n'.join(rows)


# Função principal para executar a migração
def run_migration():
    """
    Executa o processo completo de migração.
    """
    migration = GCPJDataMigration()
    
    try:
        # Carregar dados
        migration.load_data(
            gcpj_file='listagcpjnão_encontrados_octopus.csv',
            primary_file='cópiaMOYA E LARA_BASE GCPJ ATIVOS  07_04_2025.xlsx',
            secondary_file='4.MOYA E LARA SOCIEDADE DE ADVOGADOS_PRÉVIA BASE ATIVA _ABRIL_disp_24_04_2025.xlsx',
            juridico_file='BASE JURIDICO_CPJ3C  07_04_2025_dados avalistas.xlsx',
            template_file='templatebancobradescosa.xlsx'
        )
        
        # Executar migração
        migrated_data = migration.migrate_data()
        
        # Aplicar formatação
        migration.apply_data_formatting()
        
        # Gerar relatórios
        migration.generate_migration_report()
        migration.generate_quality_dashboard()
        
        # Exportar resultados
        migration.export_results()
        
        logging.info("\n✓ MIGRAÇÃO CONCLUÍDA COM SUCESSO!")
        
        return migration
        
    except Exception as e:
        logging.error(f"Erro durante a migração: {str(e)}")
        raise


if __name__ == "__main__":
    migration = run_migration()
