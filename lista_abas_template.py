from openpyxl import load_workbook

# Caminho do arquivo template com nome correto
template_path = "C:\\desenvolvimento\\migration_app\\template-banco-bradesco-sa.xlsx"

# Carregar o workbook
wb = load_workbook(template_path)

# Listar todas as abas
print("Abas do template:")
for sheet_name in wb.sheetnames:
    print(f"- {sheet_name}")

# Tentar identificar a aba principal (que contém as colunas)
main_sheet = wb[wb.sheetnames[0]]
print("\nColunas da aba principal:")
for col in range(1, main_sheet.max_column + 1):
    cell_value = main_sheet.cell(row=1, column=col).value
    if cell_value:
        print(f"{col}. {cell_value}")