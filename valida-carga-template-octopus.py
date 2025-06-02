import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from difflib import get_close_matches
import os
from datetime import datetime
import re
import numpy as np

DEBUG = True

# Variáveis globais para armazenar resultados das validações
correcoes_documentos = []
correcoes_uf = []
valores_invalidos = []
log_mensagens = []
app_log_area = None

# Mapeamentos globais para validação e revalidação
mapa_validacoes = {
    "AGÊNCIA": ("AGÊNCIAS", 4),
    "GESTOR": ("GESTOR", 0),
    "UF": ("UFs", 0),
    "MONITORAMENTO": ("MONITORAMENTO", 0),
    "PLANO": ("PLANO", 0)
}

colunas_adicionais = {
    'PROCEDIMENTO': 'PROCEDIMENTOS',
    'ORGANIZAÇÃO CLIENTE': 'ORGANIZAÇÕES_CLIENTES',
    'ESCRITÓRIO': 'ESCRITÓRIOS',
    'TIPO DE OPERAÇÃO/CARTEIRA': 'OPERAÇÕES_CARTEIRAS',
    'SEGMENTO DO CONTRATO': 'SEGMENTOS_DOS_CONTRATOS',
    'PROVIDÊNCIA': 'PROVIDÊNCIAS',
    'ÁREA RESPONSÁVEL PROCESSO (EX-DEJUR)': 'AREA_RESPONSAVEL_CLIENTE',
    'RESPONSÁVEL PROCESSO (EX-DEJUR)': 'RESPONSAVEL_PROCESSO_CLIENTE',
    'ÁREA RESPONSÁVEL PROCESSO [ESCRITÓRIO]': 'AREA_RESPONSAVEL_ESCRITORIO',
    'RESPONSÁVEL PROCESSO [ESCRITÓRIO]': 'RESPONSAVEL_PROCESSO_ESCRITORIO'
}

def log(msg):
    """Registra mensagens no log e na interface gráfica."""
    log_mensagens.append(msg)
    if DEBUG:
        print("[DEBUG]", msg)
    if app_log_area:
        app_log_area.insert(tk.END, f"{msg}\n")
        app_log_area.yview(tk.END)

def normalizar(val):
    """Normaliza um valor para comparação."""
    if pd.isna(val) or val == '':
        return ""
    return str(val).strip().upper().zfill(4) if pd.notna(val) else ""

def normalizar_lista(lista, padding=0):
    """Normaliza uma lista de valores para comparação."""
    return set(str(x).strip().upper().zfill(padding) for x in lista if pd.notna(x))

def selecionar_arquivos(titulo, multiplos=False):
    """Abre o diálogo para seleção de arquivos."""
    if multiplos:
        return filedialog.askopenfilenames(title=titulo, filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")])
    return filedialog.askopenfilename(title=titulo, filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")])

def salvar_log_txt(nome_base, timestamp, diretorio):
    """Salva o log completo em um arquivo TXT."""
    try:
        nome_log = nome_base + f"_Log_{timestamp}.txt"
        caminho_log = os.path.join(diretorio, nome_log)
        
        with open(caminho_log, 'w', encoding='utf-8') as f:
            f.write(f"LOG DE VALIDAÇÃO - {nome_base}\n")
            f.write(f"Data/Hora: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}\n")
            f.write(f"{'='*50}\n\n")
            
            for msg in log_mensagens:
                f.write(f"{msg}\n")
        
        log(f"Arquivo de log salvo: {caminho_log}")
        return caminho_log
    except Exception as e:
        log(f"Erro ao salvar arquivo de log: {str(e)}")
        return None

def remover_colunas_duplicadas_vazias(df):
    """Remove colunas duplicadas que estão vazias."""
    colunas_originais = list(df.columns)
    colunas_base = {}
    
    # Agrupar colunas pelo nome base (removendo sufixos como .1, .2)
    for col in colunas_originais:
        # Usando r"\.\d+$" para evitar problemas com o escape
        nome_base = re.sub(r"\.\d+$", "", col)
        if nome_base not in colunas_base:
            colunas_base[nome_base] = []
        colunas_base[nome_base].append(col)
    
    # Identificar duplicadas vazias para remover
    colunas_para_remover = []
    for nome_base, colunas in colunas_base.items():
        if len(colunas) > 1:
            # Manter a primeira coluna, verificar se as outras estão vazias
            for col in colunas[1:]:
                is_empty = True
                for valor in df[col]:
                    if pd.notna(valor) and str(valor).strip() != '':
                        is_empty = False
                        break
                
                if is_empty:
                    colunas_para_remover.append(col)
                    log(f"Removida coluna duplicada e vazia: {col}")
    
    # Remover as colunas identificadas
    if colunas_para_remover:
        df.drop(columns=colunas_para_remover, inplace=True)
    
    return df

def verificar_valor_na_lista(valor, lista_valida, coluna, idx, padding=0):
    """Verifica se um valor está na lista de valores válidos. 
    Não faz correções automáticas, apenas identifica valores inválidos."""
    if pd.isna(valor) or valor == '':
        return valor
        
    # Normalizar o valor e a lista para comparação
    val_normalizado = str(valor).strip().upper()
    if padding > 0:
        val_normalizado = val_normalizado.zfill(padding)
    
    # Normalizar lista se não for um conjunto
    if not isinstance(lista_valida, set):
        lista_normalizada = normalizar_lista(lista_valida, padding)
    else:
        lista_normalizada = lista_valida
    
    if val_normalizado not in lista_normalizada:
        # Tentar achar uma correspondência aproximada
        sugestao = get_close_matches(val_normalizado, lista_normalizada, n=1, cutoff=0.6)
        
        if sugestao:
            # Registrar a sugestão, mas não aplicar correção automática
            msg = f"Valor inválido encontrado na linha {idx+2}, coluna '{coluna}': '{valor}'. Sugestão: '{sugestao[0]}'"
            log(msg)
            valores_invalidos.append((coluna, idx+2, valor, sugestao[0]))
        else:
            # Registrar valor inválido sem sugestão
            msg = f"Valor inválido encontrado na linha {idx+2}, coluna '{coluna}': '{valor}' (sem sugestão)"
            log(msg)
            valores_invalidos.append((coluna, idx+2, valor, None))
        
        # Marcar como inválido
        return f"INVALIDO:{valor}"
    else:
        # Valor está correto, retornar como está
        return valor

def calcular_digito_cpf(cpf_base):
    """Calcula os dígitos verificadores de um CPF."""
    soma1 = sum(int(cpf_base[i]) * (10 - i) for i in range(9))
    d1 = (soma1 * 10 % 11) % 10
    soma2 = sum(int(cpf_base[i]) * (11 - i) for i in range(9)) + d1 * 2
    d2 = (soma2 * 10 % 11) % 10
    return cpf_base + f"{d1}{d2}"

def calcular_digito_cnpj(cnpj_base):
    """Calcula os dígitos verificadores de um CNPJ."""
    pesos1 = [5, 4, 3, 2, 9, 8, 7, 6, 5, 4, 3, 2]
    pesos2 = [6] + pesos1
    soma1 = sum(int(cnpj_base[i]) * pesos1[i] for i in range(12))
    d1 = 11 - soma1 % 11
    d1 = d1 if d1 < 10 else 0
    soma2 = sum(int(cnpj_base[i]) * pesos2[i] for i in range(12)) + d1 * pesos2[12]
    d2 = 11 - soma2 % 11
    d2 = d2 if d2 < 10 else 0
    return cnpj_base + f"{d1}{d2}"

def validar_cpf(cpf):
    """Valida um CPF."""
    if not cpf.isdigit() or len(set(cpf)) == 1:
        return False
    soma1 = sum(int(cpf[i]) * (10 - i) for i in range(9))
    d1 = (soma1 * 10 % 11) % 10
    soma2 = sum(int(cpf[i]) * (11 - i) for i in range(10))
    d2 = (soma2 * 10 % 11) % 10
    return cpf[-2:] == f"{d1}{d2}"

def validar_cnpj(cnpj):
    """Valida um CNPJ."""
    if not cnpj.isdigit() or len(set(cnpj)) == 1:
        return False
    pesos1 = [5, 4, 3, 2, 9, 8, 7, 6, 5, 4, 3, 2]
    pesos2 = [6] + pesos1
    soma1 = sum(int(cnpj[i]) * pesos1[i] for i in range(12))
    d1 = 11 - soma1 % 11
    d1 = d1 if d1 < 10 else 0
    soma2 = sum(int(cnpj[i]) * pesos2[i] for i in range(13))
    d2 = 11 - soma2 % 11
    d2 = d2 if d2 < 10 else 0
    return cnpj[-2:] == f"{d1}{d2}"

def validar_documento(doc):
    """Valida um documento (CPF ou CNPJ)."""
    doc = re.sub(r"\D", "", str(doc))
    if len(doc) == 11:
        return validar_cpf(doc)
    elif len(doc) == 14:
        return validar_cnpj(doc)
    return False

def validar_documentos(df):
    """Valida documentos (CPF/CNPJ) nas colunas correspondentes."""
    col_docs = [col for col in df.columns if "CPF" in col.upper() or "CNPJ" in col.upper()]
    log(f"Colunas de documentos encontradas: {col_docs}")
    
    for col in col_docs:
        log(f"Validando documentos na coluna: {col}")
        for i, val in df[col].items():
            if pd.isna(val) or val == '':
                continue
                
            original = str(val)
            doc = re.sub(r"\D", "", original)
            
            # Verificar se o documento tem o tamanho correto
            if len(doc) == 11 and not validar_cpf(doc):
                doc_corrigido = calcular_digito_cpf(doc[:9])
                correcoes_documentos.append((col, i + 2, original, doc_corrigido))
                df.at[i, col] = doc_corrigido
                log(f"CPF corrigido na linha {i+2}: de '{original}' para '{doc_corrigido}'")
            elif len(doc) == 14 and not validar_cnpj(doc):
                doc_corrigido = calcular_digito_cnpj(doc[:12])
                correcoes_documentos.append((col, i + 2, original, doc_corrigido))
                df.at[i, col] = doc_corrigido
                log(f"CNPJ corrigido na linha {i+2}: de '{original}' para '{doc_corrigido}'")
            elif len(doc) not in [0, 11, 14]:
                log(f"Documento inválido na linha {i+2}: '{original}' (tamanho incorreto)")
                df.at[i, col] = f"INVALIDO:{original}"

def validar_coerencia_comarca_uf(df):
    """Valida e corrige a coerência entre COMARCA e UF."""
    correcoes = []
    
    if 'COMARCA' in df.columns and 'UF' in df.columns:
        for idx, row in df.iterrows():
            comarca = str(row['COMARCA']).strip().upper()
            
            # Se a COMARCA estiver vazia, não temos como validar
            if not comarca:
                continue
                
            uf_atual = str(row['UF']).strip().upper()
            
            # Tentar extrair UF do final da COMARCA (formato "CIDADE-UF")
            match = re.search(r"-([A-Z]{2})$", comarca)
            if match:
                uf_correta = match.group(1)
                
                # Se a UF atual não corresponder à UF correta, corrigir
                if uf_atual and uf_atual != uf_correta:
                    log(f"Corrigida UF na linha {idx + 2}: de '{uf_atual}' para '{uf_correta}' (COMARCA: {comarca})")
                    correcoes.append((idx + 2, uf_atual, uf_correta))  # +2 para ajustar ao número da linha na planilha
                    df.at[idx, 'UF'] = uf_correta
    
    return correcoes

def validar_coluna_padrao(df, coluna, wb_template, aba, padding=0):
    """Valida uma coluna contra uma lista de valores em uma aba do template."""
    log(f"Validando coluna: {coluna}")
    try:
        # Obter lista de valores válidos da aba correspondente
        ws = wb_template[aba]
        lista_valida = [cell.value for cell in ws['A'] if cell.value is not None]
        lista_normalizada = normalizar_lista(lista_valida, padding)
        
        # Validar cada valor da coluna
        for idx, valor in df[coluna].items():
            if pd.isna(valor) or valor == '':
                continue
            
            # Verificar se o valor está na lista e marcar se inválido
            resultado = verificar_valor_na_lista(valor, lista_normalizada, coluna, idx, padding)
            if resultado != valor:
                df.at[idx, coluna] = resultado
    except Exception as e:
        log(f"Erro ao validar coluna {coluna}: {str(e)}")

def validar_todas_colunas_padrao(df, wb_template):
    """Valida todas as colunas padrão contra suas respectivas abas no template."""
    global mapa_validacoes
    
    # Validar cada coluna do mapeamento
    for coluna, (aba, padding) in mapa_validacoes.items():
        if coluna in df.columns and aba in wb_template.sheetnames:
            validar_coluna_padrao(df, coluna, wb_template, aba, padding)

def validar_colunas_adicionais(df, wb_template):
    """Valida colunas adicionais contra suas listas de referência."""
    global colunas_adicionais
    
    for coluna, aba in colunas_adicionais.items():
        if coluna in df.columns and aba in wb_template.sheetnames:
            validar_coluna_padrao(df, coluna, wb_template, aba)
    
    return df

def preservar_nomenclatura_exata(df_lote, df_template):
    """Garante que as colunas do lote sigam exatamente a nomenclatura do template."""
    colunas_template = list(df_template.columns)
    colunas_lote = list(df_lote.columns)
    
    # Criar mapeamento de nomes semelhantes (ignorando case, espaços extras, etc.)
    mapeamento_colunas = {}
    for col_template in colunas_template:
        col_template_norm = col_template.strip().upper()
        for col_lote in colunas_lote:
            col_lote_norm = col_lote.strip().upper()
            if col_template_norm == col_lote_norm and col_template != col_lote:
                mapeamento_colunas[col_lote] = col_template
                log(f"Coluna renomeada para padrão do template: '{col_lote}' -> '{col_template}'")
    
    # Renomear colunas conforme o mapeamento
    if mapeamento_colunas:
        df_lote = df_lote.rename(columns=mapeamento_colunas)
    
    # Adicionar colunas faltantes com nome exato do template
    for col in colunas_template:
        if col not in df_lote.columns:
            df_lote[col] = ""
            log(f"Adicionada coluna faltante: '{col}'")
    
    return df_lote

def validar_campos_obrigatorios(df):
    """Valida o preenchimento de campos obrigatórios."""
    log("Iniciando validação de campos obrigatórios...")
    
    campos_obrigatorios = [
        "PROCESSO",
        "NOME PARTE CONTRÁRIA PRINCIPAL",
        "CPF/CNPJ",
        "ORGANIZAÇÃO CLIENTE",
        "ESCRITÓRIO",
        "TIPO DE OPERAÇÃO/CARTEIRA",
        "OPERAÇÃO",
        "AGÊNCIA",
        "CONTA",
        "SEGMENTO DO CONTRATO",
        "UF"
    ]
    
    campos_vazios = {}
    
    for campo in campos_obrigatorios:
        if campo in df.columns:
            # Conta valores vazios (NaN, None, string vazia)
            vazios = df[campo].apply(lambda x: pd.isna(x) or str(x).strip() == '').sum()
            if vazios > 0:
                campos_vazios[campo] = vazios
                # Marcar células vazias como inválidas
                df.loc[df[campo].apply(lambda x: pd.isna(x) or str(x).strip() == ''), campo] = "INVALIDO:CAMPO_OBRIGATORIO"
                log(f"Campo obrigatório '{campo}' não preenchido em {vazios} linhas")
    
    return campos_vazios

def aplicar_correcoes(df, valores_invalidos, wb_template):
    """Permite ao usuário corrigir valores inválidos identificados durante a validação."""
    if not valores_invalidos:
        log("Não há valores inválidos para corrigir.")
        return df, False
    
    # Criar uma janela temporária para correções
    janela_correcao = tk.Toplevel()
    janela_correcao.title("Correção de Valores Inválidos")
    janela_correcao.geometry("900x650")
    
    # Variáveis para controlar o estado
    correcoes_aplicadas = False
    df_corrigido = df.copy()
    
    # Criar frame principal
    frame_principal = tk.Frame(janela_correcao)
    frame_principal.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
    
    # Mensagem de instruções
    tk.Label(frame_principal, text="Corrija os valores inválidos. As sugestões são pré-preenchidas quando disponíveis.", 
             font=("Arial", 10, "bold")).pack(anchor=tk.W, pady=(0, 10))
    
    # Agrupar valores inválidos por coluna para organização
    valores_por_coluna = {}
    for coluna, linha, valor, sugestao in valores_invalidos:
        if coluna not in valores_por_coluna:
            valores_por_coluna[coluna] = []
        # Calcular nível de confiança baseado na existência de sugestão
        confianca = 0.8 if sugestao else 0.0
        valores_por_coluna[coluna].append((linha, valor, sugestao, confianca))
    
    # Criar notebook para abas (uma aba por coluna)
    notebook = ttk.Notebook(frame_principal)
    notebook.pack(fill=tk.BOTH, expand=True, pady=10)
    
    # Para cada coluna, criar uma aba com sua lista de valores inválidos
    tabs = {}
    entry_widgets = {}
    check_vars = {}  # Para controle de correções em lote
    
    for coluna, valores in valores_por_coluna.items():
        # Criar frame para esta coluna
        tab = tk.Frame(notebook)
        notebook.add(tab, text=f"{coluna} ({len(valores)})")
        tabs[coluna] = tab
        
        # Frame para ferramentas de correção em lote
        batch_frame = tk.Frame(tab)
        batch_frame.pack(fill=tk.X, padx=5, pady=5)
        
        tk.Label(batch_frame, text="Correção em lote:").pack(side=tk.LEFT, padx=(0, 10))
        batch_entry = tk.Entry(batch_frame, width=30)
        batch_entry.pack(side=tk.LEFT, padx=(0, 10))
        
        def criar_apply_batch(col, entry):
            def apply_batch():
                valor_batch = entry.get().strip()
                if not valor_batch:
                    messagebox.showwarning("Aviso", "Digite um valor para aplicar em lote.")
                    return
                
                # Aplicar a todos os selecionados
                for linha, check_var in check_vars[col].items():
                    if check_var.get():
                        entry_widgets[col][linha].delete(0, tk.END)
                        entry_widgets[col][linha].insert(0, valor_batch)
                
                messagebox.showinfo("Correção em Lote", f"Valor '{valor_batch}' aplicado aos itens selecionados.")
            return apply_batch
        
        tk.Button(batch_frame, text="Aplicar aos Selecionados", 
                  command=criar_apply_batch(coluna, batch_entry)).pack(side=tk.LEFT)
        
        # Criar frame com barra de rolagem para a lista de valores
        list_frame = tk.Frame(tab)
        list_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # Adicionar barra de rolagem
        canvas = tk.Canvas(list_frame)
        scrollbar = tk.Scrollbar(list_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Criar cabeçalho
        tk.Label(scrollable_frame, text="Selecionar", width=10).grid(row=0, column=0, padx=5, pady=5)
        tk.Label(scrollable_frame, text="Linha", width=10).grid(row=0, column=1, padx=5, pady=5)
        tk.Label(scrollable_frame, text="Valor Atual", width=20).grid(row=0, column=2, padx=5, pady=5)
        tk.Label(scrollable_frame, text="Valor Correto", width=30).grid(row=0, column=3, padx=5, pady=5)
        tk.Label(scrollable_frame, text="Sugestão", width=20).grid(row=0, column=4, padx=5, pady=5)
        tk.Label(scrollable_frame, text="Confiança", width=10).grid(row=0, column=5, padx=5, pady=5)
        
        # Adicionar valores inválidos para esta coluna
        entry_widgets[coluna] = {}
        check_vars[coluna] = {}
        
        for i, (linha, valor, sugestao, confianca) in enumerate(valores):
            # Remover o prefixo "INVALIDO:" se presente
            valor_limpo = valor
            if isinstance(valor, str) and valor.startswith("INVALIDO:"):
                valor_limpo = valor[9:]
            
            # Checkbox para seleção em lote
            check_var = tk.BooleanVar(value=False)
            check_vars[coluna][linha] = check_var
            tk.Checkbutton(scrollable_frame, variable=check_var).grid(row=i+1, column=0, padx=5, pady=2)
            
            # Linha
            tk.Label(scrollable_frame, text=str(linha)).grid(row=i+1, column=1, padx=5, pady=2)
            
            # Valor atual
            tk.Label(scrollable_frame, text=valor_limpo).grid(row=i+1, column=2, padx=5, pady=2)
            
            # Campo de entrada para correção
            entry = tk.Entry(scrollable_frame, width=30)
            if sugestao:
                entry.insert(0, sugestao)
                # Se a confiança for alta, destacar em verde
                if confianca >= 0.8:
                    entry.configure(bg="#e6ffe6")  # Verde claro
            else:
                entry.insert(0, valor_limpo)
            entry.grid(row=i+1, column=3, padx=5, pady=2)
            entry_widgets[coluna][linha] = entry
            
            # Sugestão
            sugestao_text = sugestao if sugestao else "Sem sugestão"
            tk.Label(scrollable_frame, text=sugestao_text).grid(row=i+1, column=4, padx=5, pady=2)
            
            # Indicador de confiança
            confianca_text = f"{int(confianca * 100)}%" if confianca > 0 else "N/A"
            tk.Label(scrollable_frame, text=confianca_text).grid(row=i+1, column=5, padx=5, pady=2)
    
    # Frame para controles e opções
    frame_controles = tk.Frame(frame_principal)
    frame_controles.pack(fill=tk.X, pady=10)
    
    # Opção para revalidar após correções
    revalidar_var = tk.BooleanVar(value=True)
    tk.Checkbutton(frame_controles, text="Revalidar após correções", variable=revalidar_var).pack(side=tk.LEFT, padx=5)
    
    # Opção para aplicar automaticamente correções de alta confiança
    auto_aplicar_var = tk.BooleanVar(value=True)
    tk.Checkbutton(frame_controles, text="Aplicar automaticamente sugestões de alta confiança", 
                   variable=auto_aplicar_var).pack(side=tk.LEFT, padx=5)
    
    # Função para aplicar automaticamente correções de alta confiança
    def auto_aplicar():
        if not auto_aplicar_var.get():
            return
        
        aplicadas = 0
        for coluna, valores in valores_por_coluna.items():
            for linha, _, sugestao, confianca in valores:
                if sugestao and confianca >= 0.8:
                    # Linha na interface é linha+2 no Excel, mas linha na interface é linha no DataFrame (que é 0-indexed)
                    idx_df = linha - 2
                    valor_atual = df_corrigido.at[idx_df, coluna]
                    
                    if isinstance(valor_atual, str) and valor_atual.startswith("INVALIDO:"):
                        df_corrigido.at[idx_df, coluna] = sugestao
                        aplicadas += 1
                        log(f"Auto-correção aplicada: Linha {linha}, Coluna '{coluna}': '{valor_atual[9:]}' -> '{sugestao}'")
        
        if aplicadas > 0:
            messagebox.showinfo("Auto-correção", f"{aplicadas} correções de alta confiança foram aplicadas automaticamente.")
            # Fechar a janela se todas as correções foram aplicadas automaticamente
            if aplicadas == sum(len(v) for v in valores_por_coluna.values()):
                janela_correcao.destroy()
                nonlocal correcoes_aplicadas
                correcoes_aplicadas = True
    
    # Chamar auto-aplicar após um breve atraso para permitir que a interface seja renderizada
    janela_correcao.after(500, auto_aplicar)
    
    # Função para aplicar todas as correções
    def aplicar():
        nonlocal correcoes_aplicadas
        
        # Percorrer todas as entradas e aplicar correções
        correcoes_count = 0
        for coluna, entradas in entry_widgets.items():
            for linha, entry in entradas.items():
                novo_valor = entry.get().strip()
                # Linha na interface é linha+2 no Excel, mas linha na interface é linha no DataFrame (que é 0-indexed)
                idx_df = linha - 2
                valor_atual = df_corrigido.at[idx_df, coluna]
                
                # Verificar se valor atual já está marcado como inválido
                if isinstance(valor_atual, str) and valor_atual.startswith("INVALIDO:"):
                    # Se a correção não estiver vazia, aplicá-la
                    if novo_valor:
                        df_corrigido.at[idx_df, coluna] = novo_valor
                        correcoes_count += 1
                        log(f"Valor corrigido: Linha {linha}, Coluna '{coluna}': '{valor_atual[9:]}' -> '{novo_valor}'")
        
        log(f"Total de correções aplicadas: {correcoes_count}")
        correcoes_aplicadas = correcoes_count > 0
        janela_correcao.destroy()
    
    # Função para cancelar correções
    def cancelar():
        janela_correcao.destroy()
    
    # Adicionar botões
    frame_botoes = tk.Frame(frame_principal)
    frame_botoes.pack(pady=10)
    
    tk.Button(frame_botoes, text="Aplicar Correções", command=aplicar, width=20).pack(side=tk.LEFT, padx=10)
    tk.Button(frame_botoes, text="Cancelar", command=cancelar, width=20).pack(side=tk.LEFT, padx=10)
    
    # Aguardar que a janela seja fechada
    janela_correcao.wait_window()
    
    # Se opção de revalidação estiver ativada e correções foram aplicadas
    revalidar = revalidar_var.get() if 'revalidar_var' in locals() else False
    if revalidar and correcoes_aplicadas:
        log("Revalidando valores corrigidos...")
        # Revalidar colunas relevantes
        for coluna in valores_por_coluna.keys():
            # Determinar a aba correspondente à coluna
            aba = None
            # Verificar em mapa_validacoes
            for col, (tab, padding) in mapa_validacoes.items():
                if col == coluna:
                    aba, padding = tab, padding
                    validar_coluna_padrao(df_corrigido, coluna, wb_template, aba, padding)
                    break
            
            # Verificar em colunas_adicionais se não encontrado
            if not aba:
                for col, tab in colunas_adicionais.items():
                    if col == coluna:
                        aba = tab
                        validar_coluna_padrao(df_corrigido, coluna, wb_template, aba)
                        break
    
    return df_corrigido, correcoes_aplicadas

def gerar_estatisticas(df):
    """Gera estatísticas sobre as validações realizadas."""
    total_registros = len(df)
    log(f"\n== ESTATÍSTICAS ==")
    log(f"Total de registros processados: {total_registros}")
    
    # Estatísticas de correções de documentos
    if correcoes_documentos:
        log(f"Total de documentos corrigidos: {len(correcoes_documentos)}")
        for col, linha, antigo, novo in correcoes_documentos:
            log(f"  - Linha {linha}, Coluna {col}: {antigo} -> {novo}")
    else:
        log("Não foram necessárias correções de documentos.")
    
    # Estatísticas de correções de UF
    if correcoes_uf:
        log(f"Total de correções de UF: {len(correcoes_uf)}")
        for linha, antigo, novo in correcoes_uf:
            log(f"  - Linha {linha}: {antigo} -> {novo}")
    else:
        log("Não foram necessárias correções de UF por inconsistência com COMARCA.")
    
    # Estatísticas de valores inválidos
    if valores_invalidos:
        log(f"\nTotal de valores inválidos encontrados: {len(valores_invalidos)}")
        
        # Agrupar por coluna para facilitar a visualização
        colunas_com_invalidos = {}
        for coluna, linha, valor, sugestao in valores_invalidos:
            if coluna not in colunas_com_invalidos:
                colunas_com_invalidos[coluna] = 0
            colunas_com_invalidos[coluna] += 1
        
        log("\nValores inválidos por coluna:")
        for coluna, count in colunas_com_invalidos.items():
            log(f"  - {coluna}: {count} valores")
    else:
        log("\nNão foram encontrados outros valores inválidos.")
    
    # Estatísticas de campos obrigatórios não preenchidos
    if 'campos_vazios' in locals() and campos_vazios:
        log(f"\nCampos obrigatórios não preenchidos:")
        for campo, count in campos_vazios.items():
            log(f"  - {campo}: {count} linhas")
    else:
        log("\nTodos os campos obrigatórios estão preenchidos.")
    
    # Verificar se há valores marcados como inválidos
    colunas_com_invalidos = {}
    for col in df.columns:
        invalidos = df[col].astype(str).str.contains("INVALIDO:").sum()
        if invalidos > 0:
            colunas_com_invalidos[col] = invalidos
    
    if colunas_com_invalidos:
        log("\nColunas com valores marcados como inválidos:")
        for col, count in colunas_com_invalidos.items():
            log(f"  - {col}: {count} valores")
    
    log("== FIM DAS ESTATÍSTICAS ==\n")

def processar_lote(caminho_template, caminho_lote):
    """Processa um lote de dados, validando-o contra o template."""
    global correcoes_documentos, correcoes_uf, valores_invalidos
    correcoes_documentos = []
    correcoes_uf = []
    valores_invalidos = []

    log(f"Processando lote: {caminho_lote}")
    
    try:
        wb_template = load_workbook(caminho_template)
        log(f"Template carregado: {caminho_template}")
        
        df_lote = pd.read_excel(caminho_lote, sheet_name=0)
        log(f"Lote carregado: {caminho_lote} com {len(df_lote)} linhas")
        
        df_template = pd.read_excel(caminho_template, sheet_name=0)
        log(f"Template base carregado com {len(df_template.columns)} colunas")

        # Remover colunas duplicadas vazias
        df_lote = remover_colunas_duplicadas_vazias(df_lote)
        log("Verificação de colunas duplicadas concluída")

        # Preservar nomenclatura exata e adicionar colunas faltantes
        df_lote = preservar_nomenclatura_exata(df_lote, df_template)
        log("Verificação e ajuste de nomenclatura de colunas concluído")
        
        # Reordenar colunas conforme o template
        colunas_template = list(df_template.columns)
        df_lote = df_lote[colunas_template]
        log("Reorganização de colunas concluída")

        # Validar todas as colunas padrão
        log("Iniciando validação de colunas padrão...")
        validar_todas_colunas_padrao(df_lote, wb_template)

        # Validar colunas adicionais
        log("Iniciando validação de colunas adicionais...")
        df_lote = validar_colunas_adicionais(df_lote, wb_template)

        # Validar coerência entre COMARCA e UF
        log("Iniciando validação de coerência entre COMARCA e UF...")
        correcoes_uf = validar_coerencia_comarca_uf(df_lote)

        # Validar documentos (CPF/CNPJ)
        log("Iniciando validação de documentos...")
        validar_documentos(df_lote)
        
        # Validar campos obrigatórios
        log("Iniciando validação de campos obrigatórios...")
        campos_vazios = validar_campos_obrigatorios(df_lote)

        # Verificar se há valores inválidos e oferecer correção
        if valores_invalidos:
            log("Iniciando fase de correções interativas...")
            df_lote, correcoes_feitas = aplicar_correcoes(df_lote, valores_invalidos, wb_template)
            if correcoes_feitas:
                log("Correções aplicadas com sucesso!")
                # Atualizar valores_invalidos após correções
                valores_invalidos = [v for v in valores_invalidos if 
                                    not (isinstance(df_lote.at[v[1]-2, v[0]], str) and 
                                         not df_lote.at[v[1]-2, v[0]].startswith("INVALIDO:"))]

        # Gerar relatório de estatísticas
        log("Gerando estatísticas...")
        gerar_estatisticas(df_lote)

        # Salvar resultados
        log("Salvando resultado...")
        ws_principal = wb_template[wb_template.sheetnames[0]]
        ws_principal.delete_rows(2, ws_principal.max_row)
        for row in dataframe_to_rows(df_lote, index=False, header=False):
            ws_principal.append(row)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        nome_base = os.path.splitext(os.path.basename(caminho_lote))[0]
        nome_saida = nome_base + f"_Validado_{timestamp}.xlsx"
        output_path = os.path.join(os.path.dirname(caminho_lote), nome_saida)
        wb_template.save(output_path)
        log(f"Arquivo salvo: {output_path}")
        
        # Salvar relatório em formato TXT
        log_path = salvar_log_txt(nome_base, timestamp, os.path.dirname(caminho_lote))
        
        return output_path, log_path
    except Exception as e:
        log(f"ERRO CRÍTICO: {str(e)}")
        import traceback
        log(traceback.format_exc())
        raise

def iniciar_interface():
    """Inicia a interface gráfica do aplicativo."""
    global app_log_area

    root = tk.Tk()
    root.title("Validador de Lotes - Template Octopus")
    root.geometry("800x600")

    frame_principal = tk.Frame(root)
    frame_principal.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

    # Área de log com barra de rolagem
    frame_log = tk.Frame(frame_principal)
    frame_log.pack(fill=tk.BOTH, expand=True, pady=10)
    
    tk.Label(frame_log, text="Log de Processamento:").pack(anchor=tk.W)
    
    scrollbar = tk.Scrollbar(frame_log)
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
    
    app_log_area = tk.Text(frame_log, height=20, width=95, yscrollcommand=scrollbar.set)
    app_log_area.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
    scrollbar.config(command=app_log_area.yview)

    # Barra de progresso
    frame_progresso = tk.Frame(frame_principal)
    frame_progresso.pack(fill=tk.X, pady=10)
    
    tk.Label(frame_progresso, text="Progresso:").pack(anchor=tk.W)
    progresso = ttk.Progressbar(frame_progresso, orient="horizontal", length=780, mode="determinate")
    progresso.pack(fill=tk.X, pady=5)

    # Botões de ação
    frame_botoes = tk.Frame(frame_principal)
    frame_botoes.pack(pady=10)

    def executar():
        """Executa o processamento de lotes."""
        try:
            app_log_area.delete(1.0, tk.END)
            log("Iniciando processo de validação...")
            
            caminho_template = selecionar_arquivos("Selecione o arquivo Template")
            if not caminho_template:
                log("Operação cancelada: Nenhum template selecionado.")
                return
                
            log(f"Template selecionado: {caminho_template}")
            
            caminhos_lotes = selecionar_arquivos("Selecione um ou mais Lotes a Validar", multiplos=True)
            if not caminhos_lotes:
                log("Operação cancelada: Nenhum lote selecionado.")
                return
                
            log(f"Lotes selecionados: {len(caminhos_lotes)}")
            
            if isinstance(caminho_template, tuple):
                caminho_template = caminho_template[0]
                
            arquivos_salvos = []
            arquivos_log = []
            total = len(caminhos_lotes)
            progresso["maximum"] = total
            
            for i, caminho_lote in enumerate(caminhos_lotes):
                try:
                    log(f"\n{'='*40}")
                    log(f"Processando lote {i+1} de {total}: {os.path.basename(caminho_lote)}")
                    log(f"{'='*40}\n")
                    
                    saida, log_path = processar_lote(caminho_template, caminho_lote)
                    arquivos_salvos.append(saida)
                    if log_path:
                        arquivos_log.append(log_path)
                    
                    log(f"\nLote {i+1} processado com sucesso!")
                except Exception as e:
                    log(f"\nERRO ao processar {caminho_lote}: {str(e)}")
                    import traceback
                    log(traceback.format_exc())
                
                progresso["value"] = i + 1
                root.update_idletasks()

            if arquivos_salvos:
                mensagem = "Arquivos validados gerados:\n" + "\n".join(arquivos_salvos)
                
                if arquivos_log:
                    mensagem += "\n\nArquivos de log gerados:\n" + "\n".join(arquivos_log)
                
                if correcoes_documentos:
                    mensagem += f"\n\nDocumentos com dígito corrigido: {len(correcoes_documentos)}"
                
                if correcoes_uf:
                    mensagem += f"\n\nUFs corrigidas com base na COMARCA: {len(correcoes_uf)}"
                
                if valores_invalidos:
                    mensagem += f"\n\nValores inválidos encontrados: {len(valores_invalidos)}"
                
                log("\nProcessamento concluído com sucesso!")
                messagebox.showinfo("Processamento Concluído", mensagem)
            else:
                log("\nNenhum arquivo foi processado com sucesso.")
                messagebox.showwarning("Aviso", "Nenhum arquivo foi processado com sucesso. Verifique os logs para mais detalhes.")
        
        except Exception as e:
            log(f"ERRO CRÍTICO na execução: {str(e)}")
            import traceback
            log(traceback.format_exc())
            messagebox.showerror("Erro", f"Ocorreu um erro crítico: {str(e)}")

    tk.Label(frame_botoes, text="Clique para validar lotes com base no template Octopus:").pack(side=tk.LEFT, padx=5)
    tk.Button(frame_botoes, text="Selecionar e Processar Lotes", command=executar, height=2, width=30).pack(side=tk.LEFT, padx=5)

    # Versão e créditos
    tk.Label(root, text="Validador de Lotes - v2.0 | Maio/2025", fg="gray").pack(side=tk.BOTTOM, pady=5)

    root.mainloop()

if __name__ == "__main__":
    iniciar_interface()